"""
Twitter/X Post Scraper - Version 1.2.0
Scrape les réactions, retweets et commentaires d'un post X/Twitter
et exporte les données dans un fichier Excel

Améliorations v1.2.0:
- Anti-détection (User-Agent rotation, masquage WebDriver)
- Délais aléatoires pour simuler comportement humain
- Retry automatique avec gestion d'erreurs robuste
- Logging détaillé
- Détection de rate limit
- Mode headless configurable
"""

import os
import re
import sys
import time
import random
import logging
import json
from typing import Optional, List, Dict
from datetime import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configuration de l'encodage UTF-8 pour Windows
if sys.platform.startswith('win'):
    # Forcer l'encodage UTF-8 pour stdout et stderr
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr.reconfigure(encoding='utf-8')
    # Configurer la console Windows pour UTF-8
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleCP(65001)  # UTF-8 input
        kernel32.SetConsoleOutputCP(65001)  # UTF-8 output
    except:
        pass


# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Liste de User-Agents pour rotation
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
]


class TwitterScraper:
    def __init__(self, headless=False, max_retries=3):
        """Initialise le scraper avec Selenium
        
        Args:
            headless: Mode sans interface graphique
            max_retries: Nombre maximum de tentatives en cas d'erreur
        """
        self.driver = None
        self.headless = headless
        self.max_retries = max_retries
        self.stats = {
            'start_time': datetime.now(),
            'errors': [],
            'retries': 0,
            'comments_extracted': 0,
            'scroll_count': 0
        }
        logger.info("Initialisation du scraper...")
        self.setup_driver()
    
    def setup_driver(self):
        """Configure le driver Chrome avec les options nécessaires"""
        chrome_options = Options()
        
        # Mode headless si demandé
        if self.headless:
            chrome_options.add_argument('--headless=new')
            logger.info("Mode headless activé")
        
        # Options de sécurité et performance
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--lang=fr-FR')
        
        # User-Agent aléatoire pour éviter la détection
        user_agent = random.choice(USER_AGENTS)
        chrome_options.add_argument(f'--user-agent={user_agent}')
        logger.info(f"User-Agent sélectionné: {user_agent[:50]}...")
        
        # Désactiver les indicateurs d'automatisation
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Préférences pour éviter la détection
        prefs = {
            "profile.default_content_setting_values.notifications": 2,
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Masquer les propriétés webdriver
            self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                'source': '''
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                '''
            })
            
            mode = "headless" if self.headless else "visible"
            print(f"✅ Driver Chrome initialisé avec succès (mode {mode})")
            logger.info(f"Driver initialisé en mode {mode}")
            
        except Exception as e:
            logger.error(f"Erreur lors de l'initialisation du driver: {e}")
            print(f"❌ Erreur lors de l'initialisation du driver: {e}")
            raise
    
    def random_delay(self, min_seconds=1, max_seconds=3):
        """Ajoute un délai aléatoire pour simuler un comportement humain"""
        delay = random.uniform(min_seconds, max_seconds)
        time.sleep(delay)
        return delay
    
    def safe_find_element(self, by, value, timeout=10, retries=3):
        """Trouve un élément avec retry automatique
        
        Args:
            by: Type de sélecteur (By.XPATH, By.ID, etc.)
            value: Valeur du sélecteur
            timeout: Temps d'attente maximum
            retries: Nombre de tentatives
        
        Returns:
            Element trouvé ou None
        """
        for attempt in range(retries):
            try:
                element = WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((by, value))
                )
                return element
            except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
                if attempt < retries - 1:
                    logger.warning(f"Tentative {attempt + 1}/{retries} échouée pour {value}")
                    self.random_delay(1, 2)
                else:
                    logger.error(f"Élément non trouvé après {retries} tentatives: {value}")
                    return None
        return None
    
    def detect_rate_limit(self):
        """Détecte si Twitter/X a appliqué un rate limit
        
        Returns:
            bool: True si rate limit détecté
        """
        try:
            # Rechercher des indicateurs de rate limiting
            rate_limit_indicators = [
                "Rate limit exceeded",
                "Too many requests",
                "Try again later",
                "Limite de débit"
            ]
            
            page_text = self.driver.page_source.lower()
            for indicator in rate_limit_indicators:
                if indicator.lower() in page_text:
                    logger.warning(f"Rate limit détecté: {indicator}")
                    return True
            
            return False
        except Exception as e:
            logger.error(f"Erreur lors de la détection de rate limit: {e}")
            return False
    
    def extract_post_id(self, url):
        """Extrait l'ID du post depuis l'URL"""
        pattern = r'/status/(\d+)'
        match = re.search(pattern, url)
        if match:
            return match.group(1)
        return None
    
    def scrape_post_data(self, post_url):
        """Scrape les données du post (réactions, retweets, vues)"""
        print(f"\n🔍 Accès au post: {post_url}")
        logger.info(f"Début du scraping du post: {post_url}")
        
        try:
            self.driver.get(post_url)
            logger.info("Page chargée, attente du contenu...")
            
            # Délai aléatoire pour simuler un comportement humain
            delay = self.random_delay(3, 6)
            logger.info(f"Délai d'attente: {delay:.2f}s")
            
            # Vérifier le rate limit
            if self.detect_rate_limit():
                print("⚠️ Rate limit détecté! Attendez quelques minutes avant de réessayer.")
                logger.warning("Rate limit détecté lors du chargement du post")
                return None
            
            # Données du post principal
            post_data = {
                'url': post_url,
                'post_id': self.extract_post_id(post_url),
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Extraire les statistiques du post
            try:
                # Retweets
                retweet_elements = self.driver.find_elements(By.XPATH, "//button[@data-testid='retweet']//span")
                post_data['retweets'] = retweet_elements[0].text if retweet_elements else '0'
                
                # Likes
                like_elements = self.driver.find_elements(By.XPATH, "//button[@data-testid='like']//span")
                post_data['likes'] = like_elements[0].text if like_elements else '0'
                
                # Réponses (commentaires)
                reply_elements = self.driver.find_elements(By.XPATH, "//button[@data-testid='reply']//span")
                post_data['replies'] = reply_elements[0].text if reply_elements else '0'
                
                # Vues
                view_elements = self.driver.find_elements(By.XPATH, "//a[contains(@href, '/analytics')]//span")
                post_data['views'] = view_elements[0].text if view_elements else '0'
                
                print(f"✅ Statistiques du post extraites:")
                print(f"   • Retweets: {post_data['retweets']}")
                print(f"   • Likes: {post_data['likes']}")
                print(f"   • Réponses: {post_data['replies']}")
                print(f"   • Vues: {post_data['views']}")
                
                logger.info(f"Statistiques extraites: RT={post_data['retweets']}, Likes={post_data['likes']}, Réponses={post_data['replies']}")
                
            except Exception as e:
                logger.error(f"Erreur lors de l'extraction des statistiques: {e}")
                print(f"⚠️ Erreur lors de l'extraction des statistiques: {e}")
            
            return post_data
            
        except Exception as e:
            logger.error(f"Erreur lors du scraping du post: {e}")
            print(f"❌ Erreur lors du scraping: {e}")
            self.stats['errors'].append(str(e))
            return None
    
    def scrape_comments(self, max_comments=None):
        """Scrape les commentaires du post (tous par défaut)"""
        print(f"\n💬 Extraction des commentaires...")
        print("⏳ Chargement initial de la page...")
        logger.info("Début de l'extraction des commentaires")
        comments = []
        seen_tweets = set()  # Pour éviter les doublons
        
        try:
            # Attendre plus longtemps pour le chargement initial
            delay = self.random_delay(4, 7)
            logger.info(f"Attente initiale: {delay:.2f}s")
            print("   ⏳ Attente du chargement des commentaires...")
            
            # Faire quelques scrolls forcés pour déclencher le chargement
            print("   🔄 Scrolls forcés pour charger les commentaires...")
            for i in range(3):
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                self.random_delay(2, 4)
                current_tweets = len(self.driver.find_elements(By.XPATH, "//article[@data-testid='tweet']"))
                print(f"   📊 Scroll forcé #{i+1} - {current_tweets} tweets détectés")
                logger.info(f"Scroll forcé #{i+1}: {current_tweets} tweets")
            
            # Remonter en haut pour commencer l'extraction
            self.driver.execute_script("window.scrollTo(0, 0);")
            self.random_delay(1, 2)
            
            print("\n⏳ Défilement pour charger tous les commentaires disponibles...")
            
            # Scroll continu pour charger TOUS les commentaires
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            scroll_attempts = 0
            no_new_content_count = 0
            max_no_new_content = 5
            
            while no_new_content_count < max_no_new_content:
                # Scroll vers le bas avec délai aléatoire
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                delay = self.random_delay(2, 4)
                
                # Calculer la nouvelle hauteur
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                
                # Compter les tweets actuellement visibles
                current_tweets = len(self.driver.find_elements(By.XPATH, "//article[@data-testid='tweet']"))
                
                if new_height == last_height:
                    no_new_content_count += 1
                    print(f"   📊 Tentative {no_new_content_count}/{max_no_new_content} - {current_tweets} tweets chargés")
                    logger.info(f"Pas de nouveau contenu: tentative {no_new_content_count}/{max_no_new_content}")
                else:
                    no_new_content_count = 0
                    scroll_attempts += 1
                    print(f"   📊 Scroll #{scroll_attempts} - {current_tweets} tweets chargés")
                    logger.info(f"Scroll #{scroll_attempts}: {current_tweets} tweets")
                    self.stats['scroll_count'] = scroll_attempts
                    
                last_height = new_height
            
            print(f"\n✅ Défilement terminé après {scroll_attempts} scrolls")
            print("📝 Extraction des données des commentaires...")
            logger.info(f"Défilement terminé: {scroll_attempts} scrolls effectués")
            
            # Remonter en haut avant d'extraire
            self.driver.execute_script("window.scrollTo(0, 0);")
            self.random_delay(1, 2)
            
            # Extraire les commentaires
            tweet_elements = self.driver.find_elements(By.XPATH, "//article[@data-testid='tweet']")
            total_tweets = len(tweet_elements)
            
            print(f"   Total de tweets trouvés: {total_tweets}")
            logger.info(f"Total de tweets trouvés: {total_tweets}")
            
            if total_tweets <= 1:
                print("   ⚠️ ATTENTION: Aucun commentaire détecté!")
                print("   💡 Suggestions:")
                print("      - Le post n'a peut-être pas de commentaires")
                print("      - Twitter/X a peut-être bloqué le chargement")
                print("      - Essayez de désactiver le mode headless")
                logger.warning("Aucun commentaire détecté")
                return comments
            
            for idx, tweet in enumerate(tweet_elements[1:]):  # Skip le premier (post original)
                try:
                    # Créer un identifiant unique pour éviter les doublons
                    tweet_html = tweet.get_attribute('outerHTML')[:200]
                    
                    if tweet_html in seen_tweets:
                        continue
                    
                    seen_tweets.add(tweet_html)
                    
                    comment_data = {}
                    
                    # Nom d'utilisateur
                    try:
                        username_elem = tweet.find_element(By.XPATH, ".//div[@data-testid='User-Name']")
                        comment_data['username'] = username_elem.text.split('\n')[0]
                    except:
                        comment_data['username'] = 'N/A'
                    
                    # Handle (@username)
                    try:
                        handle_elem = tweet.find_element(By.XPATH, ".//div[@data-testid='User-Name']//span[contains(text(), '@')]")
                        comment_data['handle'] = handle_elem.text
                    except:
                        comment_data['handle'] = 'N/A'
                    
                    # Texte du commentaire
                    try:
                        text_elem = tweet.find_element(By.XPATH, ".//div[@data-testid='tweetText']")
                        comment_data['text'] = text_elem.text
                    except:
                        comment_data['text'] = 'N/A'
                    
                    # Timestamp
                    try:
                        time_elem = tweet.find_element(By.XPATH, ".//time")
                        comment_data['timestamp'] = time_elem.get_attribute('datetime')
                    except:
                        comment_data['timestamp'] = 'N/A'
                    
                    # Likes du commentaire
                    try:
                        like_elem = tweet.find_element(By.XPATH, ".//button[@data-testid='like']//span")
                        comment_data['likes'] = like_elem.text if like_elem.text else '0'
                    except:
                        comment_data['likes'] = '0'
                    
                    # Retweets du commentaire
                    try:
                        retweet_elem = tweet.find_element(By.XPATH, ".//button[@data-testid='retweet']//span")
                        comment_data['retweets'] = retweet_elem.text if retweet_elem.text else '0'
                    except:
                        comment_data['retweets'] = '0'
                    
                    # Réponses du commentaire
                    try:
                        reply_elem = tweet.find_element(By.XPATH, ".//button[@data-testid='reply']//span")
                        comment_data['replies'] = reply_elem.text if reply_elem.text else '0'
                    except:
                        comment_data['replies'] = '0'
                    
                    comments.append(comment_data)
                    self.stats['comments_extracted'] = len(comments)
                    
                    # Afficher la progression tous les 10 commentaires
                    if (idx + 1) % 10 == 0:
                        print(f"   ⏳ {len(comments)} commentaires extraits...")
                        logger.info(f"{len(comments)} commentaires extraits")
                    
                    # Si max_comments est défini, s'arrêter à cette limite
                    if max_comments and len(comments) >= max_comments:
                        print(f"   ⚠️ Limite de {max_comments} commentaires atteinte")
                        logger.info(f"Limite de {max_comments} commentaires atteinte")
                        break
                    
                except Exception as e:
                    # Ignorer silencieusement les erreurs mineures
                    logger.debug(f"Erreur mineure lors de l'extraction d'un commentaire: {e}")
                    continue
            
            print(f"\n✅ {len(comments)} commentaires uniques extraits avec succès!")
            logger.info(f"Extraction terminée: {len(comments)} commentaires")
            return comments
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction des commentaires: {e}")
            print(f"❌ Erreur lors de l'extraction des commentaires: {e}")
            import traceback
            traceback.print_exc()
            self.stats['errors'].append(str(e))
            return comments
    
    def export_to_excel(self, post_data, comments, filename=None):
        """Exporte les données vers un fichier Excel"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"twitter_scrape_{timestamp}.xlsx"
        
        print(f"\n📝 Export vers Excel: {filename}")
        logger.info(f"Début de l'export vers {filename}")
        
        try:
            # Créer un workbook
            wb = Workbook()
            
            # Feuille 1: Statistiques du post
            ws_stats = wb.active
            ws_stats.title = "Statistiques Post"
            
            # En-têtes avec style
            headers_stats = ['Métrique', 'Valeur']
            ws_stats.append(headers_stats)
            
            # Style des en-têtes
            for cell in ws_stats[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Données du post
            ws_stats.append(['URL', post_data.get('url', 'N/A')])
            ws_stats.append(['Post ID', post_data.get('post_id', 'N/A')])
            ws_stats.append(['Date de scraping', post_data.get('timestamp', 'N/A')])
            ws_stats.append(['Retweets', post_data.get('retweets', '0')])
            ws_stats.append(['Likes', post_data.get('likes', '0')])
            ws_stats.append(['Réponses', post_data.get('replies', '0')])
            ws_stats.append(['Vues', post_data.get('views', '0')])
            ws_stats.append(['Commentaires extraits', str(len(comments))])
            ws_stats.append(['Scrolls effectués', str(self.stats.get('scroll_count', 0))])
            
            # Ajuster la largeur des colonnes
            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 50
            
            # Feuille 2: Commentaires
            if comments:
                ws_comments = wb.create_sheet(title="Commentaires")
                
                # En-têtes
                headers_comments = ['#', 'Nom d\'utilisateur', 'Handle', 'Texte', 'Date', 'Likes', 'Retweets', 'Réponses']
                ws_comments.append(headers_comments)
                
                # Style des en-têtes
                for cell in ws_comments[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Données des commentaires
                for idx, comment in enumerate(comments, 1):
                    ws_comments.append([
                        idx,
                        comment.get('username', 'N/A'),
                        comment.get('handle', 'N/A'),
                        comment.get('text', 'N/A'),
                        comment.get('timestamp', 'N/A'),
                        comment.get('likes', '0'),
                        comment.get('retweets', '0'),
                        comment.get('replies', '0')
                    ])
                
                # Ajuster la largeur des colonnes
                ws_comments.column_dimensions['A'].width = 5
                ws_comments.column_dimensions['B'].width = 20
                ws_comments.column_dimensions['C'].width = 20
                ws_comments.column_dimensions['D'].width = 60
                ws_comments.column_dimensions['E'].width = 20
                ws_comments.column_dimensions['F'].width = 10
                ws_comments.column_dimensions['G'].width = 10
                ws_comments.column_dimensions['H'].width = 10
            
            # Sauvegarder le fichier
            wb.save(filename)
            print(f"✅ Fichier Excel créé avec succès: {filename}")
            logger.info(f"Export réussi: {filename}")
            
            return filename
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {e}")
            print(f"❌ Erreur lors de l'export Excel: {e}")
            return None
    
    def get_stats(self):
        """Retourne les statistiques de la session"""
        self.stats['end_time'] = datetime.now()
        self.stats['duration'] = (self.stats['end_time'] - self.stats['start_time']).total_seconds()
        return self.stats
    
    def close(self):
        """Ferme le driver"""
        if self.driver:
            self.driver.quit()
            print("\n🔒 Driver fermé")
            logger.info("Driver fermé")
            
            # Afficher les statistiques finales
            stats = self.get_stats()
            logger.info(f"Statistiques de session: {stats}")


def main():
    """Fonction principale"""
    print("=" * 60)
    print("🐦 TWITTER/X POST SCRAPER v1.2.0")
    print("=" * 60)
    
    # Demander l'URL du post
    post_url = input("\n📎 Entrez l'URL du post Twitter/X: ").strip()
    
    if not post_url:
        print("❌ URL invalide!")
        return
    
    # Demander le nombre de commentaires à extraire
    print("\n💬 Nombre de commentaires à extraire:")
    print("   1. Tous les commentaires (recommandé)")
    print("   2. Nombre limité")
    choice = input("\nVotre choix (1 ou 2): ").strip()
    
    max_comments = None
    if choice == "2":
        try:
            max_comments = int(input("Entrez le nombre maximum de commentaires: ").strip())
            print(f"✅ Limite définie à {max_comments} commentaires")
        except ValueError:
            print("⚠️ Valeur invalide, extraction de tous les commentaires par défaut")
            max_comments = None
    else:
        print("✅ Extraction de TOUS les commentaires disponibles")
    
    # Demander le mode d'affichage
    print("\n👁️ Mode d'affichage:")
    print("   1. Mode visible (recommandé pour debug)")
    print("   2. Mode headless (invisible)")
    mode_choice = input("\nVotre choix (1 ou 2): ").strip()
    headless = (mode_choice == "2")
    
    # Initialiser le scraper
    scraper = TwitterScraper(headless=headless)
    
    try:
        # Scraper les données du post
        print("\n" + "=" * 60)
        print("📊 ÉTAPE 1: Extraction des données du post")
        print("=" * 60)
        
        post_data = scraper.scrape_post_data(post_url)
        
        if not post_data:
            print("\n❌ Échec de l'extraction des données du post")
            return
        
        # Scraper les commentaires
        print("\n" + "=" * 60)
        print("📊 ÉTAPE 2: Extraction des commentaires")
        print("=" * 60)
        
        comments = scraper.scrape_comments(max_comments=max_comments)
        
        # Exporter vers Excel
        print("\n" + "=" * 60)
        print("📊 ÉTAPE 3: Export vers Excel")
        print("=" * 60)
        
        filename = scraper.export_to_excel(post_data, comments)
        
        if filename:
            print(f"\n{'=' * 60}")
            print(f"✅ SCRAPING TERMINÉ AVEC SUCCÈS!")
            print(f"{'=' * 60}")
            print(f"📁 Fichier: {filename}")
            print(f"📊 Statistiques du post:")
            print(f"   • Retweets: {post_data.get('retweets', '0')}")
            print(f"   • Likes: {post_data.get('likes', '0')}")
            print(f"   • Vues: {post_data.get('views', '0')}")
            print(f"💬 Commentaires extraits: {len(comments)}")
            
            # Afficher les statistiques de session
            stats = scraper.get_stats()
            print(f"⏱️ Durée totale: {stats['duration']:.2f}s")
            print(f"🔄 Scrolls effectués: {stats['scroll_count']}")
            if stats['errors']:
                print(f"⚠️ Erreurs rencontrées: {len(stats['errors'])}")
            print(f"{'=' * 60}")
        else:
            print("\n❌ Échec de l'export")
    
    except Exception as e:
        print(f"\n❌ Erreur: {e}")
        logger.error(f"Erreur dans main: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Fermer le driver
        print("\n🔒 Fermeture du navigateur...")
        scraper.close()


if __name__ == "__main__":
    main()
